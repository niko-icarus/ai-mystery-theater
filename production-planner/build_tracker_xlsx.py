#!/usr/bin/env python3
"""
The Lineup — Asset Tracker Spreadsheet Builder
Generates an Excel workbook with:
  Tab 1: Reusable Still Images
  Tab 2: Reusable Video Clips
  Tab 3-10: Per-episode non-reusable assets (Ep 1-8)

Each row has a "Status" column for Anthony to mark off as completed.
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
SEASONS_DIR = BASE_DIR / "seasons"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2D2D44", end_color="2D2D44", fill_type="solid")
SUBHEADER_FONT = Font(name="Inter", bold=True, size=10, color="FFD700")
DATA_FONT = Font(name="Inter", size=10)
PROMPT_FONT = Font(name="Consolas", size=9, color="8B7FC7")
STATUS_FONT = Font(name="Inter", size=10, bold=True)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="333355"),
)
GOLD_FONT = Font(name="Inter", bold=True, size=10, color="D4A020")

STATUS_OPTIONS = ["Not Started", "In Progress", "Generated", "Approved", "Final"]

# Model colors for reference
MODEL_COLORS = {
    "Claude": "D4A020",
    "ChatGPT": "32B850",
    "Gemini": "4688E8",
    "Mistral": "20C4C4",
    "DeepSeek": "D04040",
    "Llama": "8C40C0",
    "Qwen": "D4B830",
    "Grok": "A0A0A8",
}

STYLE_DNA = "layered cut paper, angular construction, visible paper edges and depth, Tim Burton papercraft aesthetic, gothic mood"

MJ_SCENE = ""
MJ_CHAR = ""
MJ_PROP = ""
MJ_SQUARE = ""
VID_SUFFIX = ""


def styled_prompt(base: str, suffix: str = "") -> str:
    """Prepend the Tim Burton papercraft style DNA to every prompt."""
    parts = [STYLE_DNA, base.rstrip(", ")]
    if suffix:
        parts.append(suffix)
    return ", ".join(p for p in parts if p)

CHARACTER_STATES = [
    ("speaking", "speaking calmly, mid-conversation, neutral expression"),
    ("defensive", "defensive posture, arms slightly raised, guarded expression"),
    ("nervous", "visibly anxious, fidgeting, avoiding eye contact"),
    ("accusatory", "pointing or gesturing emphatically, accusatory stance, intense eyes"),
    ("contemplative", "deep in thought, hand near chin, introspective, calculating gaze"),
    ("confident", "composed and confident, slight smirk, relaxed shoulders"),
    ("shocked", "wide eyes, mouth slightly open, genuine surprise"),
    ("angry", "visible fury, clenched fists, narrowed eyes, barely contained rage"),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_subheader_row(ws, row, num_cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")


def add_data_row(ws, row, values, is_prompt_col=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        if is_prompt_col and col == is_prompt_col:
            cell.font = PROMPT_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        elif col == 1:  # Status column
            cell.font = STATUS_FONT
        else:
            cell.font = DATA_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_validation(ws, col_letter, min_row, max_row):
    """Add dropdown validation for status column."""
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Generated,Approved,Final"',
        allow_blank=True,
    )
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build_workbook(story_config: dict, rotation: list, output_path: str):
    wb = Workbook()
    
    loc = story_config.get("location", {})
    loc_name = loc.get("name", "")
    loc_desc = loc.get("description", "")
    period = "1890s Victorian colonial era"
    atmosphere = "torchlit underground, dusty stone, ancient Egyptian"
    victim = story_config.get("victim", {})
    suspects = story_config.get("suspects", [])
    
    detective_name = "Inspector William Cross"
    detective_desc = "Scotland Yard detective, 48, methodical, sharp-eyed, period suit"
    
    # =================================================================
    # TAB 1: REUSABLE STILL IMAGES
    # =================================================================
    ws1 = wb.active
    ws1.title = "Reusable Stills"
    ws1.sheet_properties.tabColor = "D4A020"
    
    headers = ["Status", "Asset ID", "Category", "Name", "Description", "Reuse", "Variants", "Midjourney Prompt"]
    ws1.append(headers)
    style_header_row(ws1, 1, len(headers))
    set_col_widths(ws1, [14, 30, 16, 30, 40, 16, 20, 80])
    
    row = 2
    
    # --- Character Portraits ---
    style_subheader_row(ws1, row, len(headers), "🎭 CHARACTER PORTRAITS — Suspects")
    row += 1
    
    for s in suspects:
        for state, desc in CHARACTER_STATES:
            add_data_row(ws1, row, [
                "Not Started",
                f"char_{s['name'].lower().replace(' ', '_')}_{state}",
                "Character",
                f"{s['name']} — {state.title()}",
                f"{s['role']}. {desc}",
                "All 8 episodes",
                "Keep all 4 MJ variations",
                f"{STYLE_DNA}, {s['name']}, {s['role']}, {s['background'][:80]}, {desc}, dramatic portrait lighting, {atmosphere}, {period}",
            ], is_prompt_col=8)
            row += 1
    
    # Detective
    style_subheader_row(ws1, row, len(headers), "🔍 CHARACTER PORTRAITS — Detective")
    row += 1
    
    for state, desc in CHARACTER_STATES:
        add_data_row(ws1, row, [
            "Not Started",
            f"char_detective_{state}",
            "Character",
            f"{detective_name} — {state.title()}",
            f"{detective_desc}. {desc}",
            "All 8 episodes",
            "Keep all 4 MJ variations",
            f"{STYLE_DNA}, {detective_name}, {detective_desc}, {desc}, dramatic portrait lighting, {atmosphere}, {period}",
        ], is_prompt_col=8)
        row += 1
    
    # Victim
    style_subheader_row(ws1, row, len(headers), "💀 CHARACTER PORTRAITS — Victim")
    row += 1
    
    add_data_row(ws1, row, [
        "Not Started", "char_victim_alive", "Character",
        f"{victim['name']} — Alive",
        f"{victim['description'][:100]}",
        "All 8 episodes", "Keep all 4",
        f"{STYLE_DNA}, {victim['name']}, {victim['description'][:100]}, alive, rough demeanor, suspicious, dramatic portrait, {atmosphere}, {period}",
    ], is_prompt_col=8)
    row += 1
    
    add_data_row(ws1, row, [
        "Not Started", "char_victim_dead", "Character",
        f"{victim['name']} — Crime Scene",
        f"Deceased. {victim['cause_of_death'][:80]}",
        "All 8 episodes", "Keep best 1-2",
        f"{STYLE_DNA}, Crime scene, {victim['name']} {victim['cause_of_death'][:60]}, dramatic overhead angle, {atmosphere}, {period}, dark forensic",
    ], is_prompt_col=8)
    row += 1
    
    # --- Setting Shots ---
    style_subheader_row(ws1, row, len(headers), "🏛️ SETTING SHOTS")
    row += 1
    
    settings = [
        ("set_burial_chamber_wide", "Burial Chamber — Wide", "Full chamber, golden sarcophagus center, torches on walls, painted hieroglyphs"),
        ("set_burial_chamber_sarcophagus", "Burial Chamber — Sarcophagus Detail", "Close-up of golden sarcophagus, intricate carvings, torchlight on gold"),
        ("set_burial_chamber_hieroglyphs", "Burial Chamber — Hieroglyphs", "Wall of painted hieroglyphs, afterlife scenes, torchlight casting shadows"),
        ("set_burial_chamber_crime_scene", "Burial Chamber — Crime Scene", "Area beside sarcophagus where body was found, dark stain, scattered equipment"),
        ("set_antechamber_wide", "Antechamber — Wide", "Expedition equipment, crates and tools, torch brackets on stone walls"),
        ("set_antechamber_equipment", "Antechamber — Equipment Pile", "Excavation knives and chisels lined up, broad blades, dusty handles"),
        ("set_antechamber_carvings", "Antechamber — Wall Carvings", "Stone walls with carved inscriptions, lantern light, notes pinned nearby"),
        ("set_upper_corridor_wide", "Upper Corridor — Wide", "Long corridor into darkness, torch brackets at intervals, sand on floor"),
        ("set_upper_corridor_sealed", "Upper Corridor — Sealed Entrance", "Sealed entrance shaft, sand piled high blocking exit"),
        ("set_upper_corridor_shadow", "Upper Corridor — Shadow Figure", "Shadowy figure at far end, silhouette against torchlight"),
        ("set_lower_passage_wide", "Lower Passage — Wide", "Narrow stone passage descending, oil lamp, hieroglyphs on both walls"),
        ("set_lower_passage_conversation", "Lower Passage — Conversation Spot", "Wider section where two people talk, intimate and secretive"),
        ("set_second_level_wide", "Second Level — Wide", "Unexplored level, rough-hewn stone, single oil lamp, mysterious"),
        ("set_second_level_mapping", "Second Level — Mapping Station", "Cartographer's instruments, detailed drawings pinned to board"),
        ("set_second_level_hidden", "Second Level — Hidden Passage", "Hidden connecting passage, narrow crawlspace, dramatic perspective"),
        ("set_exterior_storm", "Exterior — Sandstorm", "Valley of the Kings, violent sandstorm, tomb entrance barely visible"),
        ("set_exterior_camp", "Exterior — Expedition Camp", "Tents and equipment battered by sand, lanterns swinging"),
        ("set_exterior_entrance", "Exterior — Entrance Shaft", "Looking down into dark entrance shaft, sand pouring in"),
    ]
    
    for sid, sname, sdesc in settings:
        add_data_row(ws1, row, [
            "Not Started", sid, "Setting", sname, sdesc,
            "All 8 episodes", "Keep all 4 MJ variations",
            f"{STYLE_DNA}, {sdesc}, inside ancient Egyptian tomb, {atmosphere}, {period}, dramatic shadows",
        ], is_prompt_col=8)
        row += 1
    
    # --- Props & Evidence ---
    style_subheader_row(ws1, row, len(headers), "🔍 PROPS & EVIDENCE")
    row += 1
    
    props = [
        ("prop_excavation_knife", "Murder Weapon — Excavation Knife", "Heavy broad-bladed excavation knife, faint dark residue in handle rivets, blade recently cleaned"),
        ("prop_equipment_pile", "Equipment Pile — Excavation Tools", "Array of excavation tools — knives, chisels, picks, one knife subtly cleaner"),
        ("prop_threatening_note", "Threatening Note", "Crumpled note, block capitals: 'THIS IS YOUR LAST WARNING'"),
        ("prop_blackmail_ledger", "Blackmail Ledger", "Hidden list of names and amounts, Victorian handwriting, pound sterling"),
        ("prop_torn_fabric", "Torn Fabric on Torch Bracket", "Dark coat lining scrap caught on iron torch bracket, stone wall"),
        ("prop_boot_prints", "Boot Prints in Dust", "Two sets in tomb dust — heavy work boots + lighter set approaching from behind"),
        ("prop_brandy_glasses", "Two Brandy Glasses", "Two glasses beside decanter, one used, one suspiciously wiped clean"),
        ("prop_hollowed_bible", "Hollowed Bible", "Open Bible with hollowed center containing lockpicks and derringer pistol"),
        ("prop_torn_notebook", "Notebook with Torn Pages", "Academic notebook, ragged torn edges, hieroglyphic sketches, ink stains"),
        ("prop_hidden_camera", "Hidden Camera Equipment", "1890s concealed camera, brass and leather, wrapped in cloth among mapping tools"),
        ("prop_revolver", "Recently Fired Revolver", "Victorian military revolver, recently fired, holster and belt nearby"),
        ("prop_sarcophagus_sketches", "Sarcophagus Mechanism Sketches", "Notebook with detailed technical diagrams of sarcophagus opening mechanism"),
        ("prop_gold_coin", "Gold Coin — Advance Payment", "Single gold coin on stone floor near tomb entrance, Spanish gold"),
    ]
    
    for pid, pname, pdesc in props:
        add_data_row(ws1, row, [
            "Not Started", pid, "Prop/Evidence", pname, pdesc,
            "All 8 episodes", "Keep all 4 MJ variations",
            f"{STYLE_DNA}, {pdesc}, dramatic overhead spotlight, dark surface, forensic photography, {period}",
        ], is_prompt_col=8)
        row += 1
    
    # --- Model Avatars ---
    style_subheader_row(ws1, row, len(headers), "🤖 MODEL AVATARS (Permanent)")
    row += 1
    
    models = ["Claude", "ChatGPT", "Gemini", "Mistral", "DeepSeek", "Llama", "Qwen", "Grok"]
    symbols = ["☀️", "⬡", "◇", "🌀", "🔺", "🦙", "✦", "✕"]
    
    for m, sym in zip(models, symbols):
        color = MODEL_COLORS[m]
        add_data_row(ws1, row, [
            "Not Started", f"avatar_{m.lower()}", "Model Avatar",
            f"{m} {sym}", f"Abstract AI avatar, color: #{color}",
            "ALL seasons", "Keep best 1-2",
            f"{STYLE_DNA}, Abstract AI avatar representing {m}, color palette centered on #{color}, contemplative digital entity, dark background, mysterious, glowing accents",
        ], is_prompt_col=8)
        row += 1
    
    # Add status dropdown validation
    add_validation(ws1, "A", 2, row)
    
    # Freeze header
    ws1.freeze_panes = "A2"
    
    # =================================================================
    # TAB 2: REUSABLE VIDEO CLIPS
    # =================================================================
    ws2 = wb.create_sheet("Reusable Videos")
    ws2.sheet_properties.tabColor = "4688E8"
    
    vid_headers = ["Status", "Asset ID", "Category", "Name", "Description", "Reuse", "Variants", "Video Gen Prompt"]
    ws2.append(vid_headers)
    style_header_row(ws2, 1, len(vid_headers))
    set_col_widths(ws2, [14, 30, 16, 30, 40, 16, 20, 80])
    
    row = 2
    
    # --- Setting Video Clips ---
    style_subheader_row(ws2, row, len(vid_headers), "🏛️ SETTING & ATMOSPHERE CLIPS")
    row += 1
    
    setting_vids = [
        ("vid_exterior_sandstorm", "Exterior — Sandstorm Over Valley", "Aerial/wide of Valley of the Kings, violent sandstorm, dramatic clouds, tomb entrance below"),
        ("vid_descent_into_tomb", "Descent — Entering the Tomb", "POV/tracking descending into tomb, light fading, torches appearing, hieroglyphs emerging"),
        ("vid_burial_chamber_pan", "Burial Chamber — Slow Pan", "Slow cinematic pan across burial chamber, sarcophagus, painted walls, torchlight"),
        ("vid_corridor_torchlight", "Corridor — Torch Flicker Walk", "Tracking through stone corridor, torches flickering, long shadows, claustrophobic"),
        ("vid_sarcophagus_reveal", "Sarcophagus — Dramatic Reveal", "Slow push-in on golden sarcophagus, torchlight catching gold, dust particles"),
        ("vid_sand_falling", "Atmosphere — Sand Through Cracks", "Close-up sand streaming through ceiling cracks, tomb being sealed, urgency"),
        ("vid_hieroglyphs_firelight", "Atmosphere — Hieroglyphs in Firelight", "Slow pan across hieroglyphs, firelight dancing, afterlife scenes in shadow"),
        ("vid_torch_flicker", "Atmosphere — Torch Close-Up", "Extreme close-up torch flame guttering, sparks, smoke, shadows leaping"),
        ("vid_evidence_table", "Evidence — Items on Surface", "Slow overhead tracking of evidence laid out — knife, note, fabric, boot print cast"),
        ("vid_sealed_entrance", "Sealed Entrance — Sand Collapse", "Entrance shaft collapsing with sand, dust cloud billowing inward, trapped"),
        ("vid_group_gathered", "Group — Suspects Assembled", "Wide shot of Victorian explorers gathered in burial chamber, torchlight, tension"),
        ("vid_transition_darkness", "Transition — Into Darkness", "Fade into deep darkness, single torch receding, transition beat"),
    ]
    
    for vid, vname, vdesc in setting_vids:
        add_data_row(ws2, row, [
            "Not Started", vid, "Setting", vname, vdesc,
            "All 8 episodes", "2-3 takes, keep all",
            f"{STYLE_DNA}, {vdesc}, ancient Egyptian tomb, {period}, cinematic camera movement, dramatic lighting",
        ], is_prompt_col=8)
        row += 1
    
    # --- Character Video Clips ---
    style_subheader_row(ws2, row, len(vid_headers), "🎭 CHARACTER ACTION CLIPS — Suspects")
    row += 1
    
    char_actions = [
        ("speaking_medium", "Speaking — Medium Shot", "speaking in medium shot, subtle gestures, conversation"),
        ("listening_closeup", "Listening — Close-Up", "listening intently, close-up, slight reactions, watchful eyes"),
        ("turning_away", "Turning Away", "turning away, walking into shadow, evasive body language"),
        ("confrontation", "Confrontation — Two Shot", "two-shot confrontation, facing another person, tense exchange"),
    ]
    
    for s in suspects:
        for action_id, action_name, action_desc in char_actions:
            add_data_row(ws2, row, [
                "Not Started",
                f"vid_{s['name'].lower().replace(' ', '_')}_{action_id}",
                "Character",
                f"{s['name']} — {action_name}",
                f"{s['role']}. {action_desc}",
                "All 8 episodes", "2-3 takes, keep all",
                f"{STYLE_DNA}, {s['name']}, {s['role']}, {action_desc}, inside {loc_name}, {atmosphere}, {period}, dramatic lighting",
            ], is_prompt_col=8)
            row += 1
    
    # Detective video clips
    style_subheader_row(ws2, row, len(vid_headers), "🔍 CHARACTER ACTION CLIPS — Detective")
    row += 1
    
    det_actions = [
        ("examining_evidence", "Examining Evidence", "examining evidence closely, turning object in hands, analytical"),
        ("pacing_thinking", "Pacing & Thinking", "pacing slowly, deep in thought, silhouette against torchlight"),
        ("interrogating", "Interrogating", "leaning forward, questioning suspect, intense eye contact"),
        ("dramatic_accusation", "Dramatic Accusation", "pointing accusingly, dramatic gesture, spotlight moment"),
    ]
    
    for action_id, action_name, action_desc in det_actions:
        add_data_row(ws2, row, [
            "Not Started",
            f"vid_detective_{action_id}",
            "Character",
            f"{detective_name} — {action_name}",
            f"{detective_desc}. {action_desc}",
            "All 8 episodes", "2-3 takes, keep all",
            f"{STYLE_DNA}, {detective_name}, {detective_desc}, {action_desc}, inside {loc_name}, {atmosphere}, {period}, dramatic lighting",
        ], is_prompt_col=8)
        row += 1
    
    add_validation(ws2, "A", 2, row)
    ws2.freeze_panes = "A2"
    
    # =================================================================
    # TABS 3-10: PER-EPISODE NON-REUSABLE ASSETS
    # =================================================================
    
    for ep_data in rotation:
        ep_num = ep_data["episode"]
        ws = wb.create_sheet(f"Episode {ep_num}")
        ws.sheet_properties.tabColor = "D04040"
        
        ep_headers = ["Status", "Asset ID", "Category", "Name", "Description", "Notes"]
        ws.append(ep_headers)
        style_header_row(ws, 1, len(ep_headers))
        set_col_widths(ws, [14, 30, 16, 30, 50, 50])
        
        row = 2
        
        # Episode info header
        det = ep_data["detective"]
        style_subheader_row(ws, row, len(ep_headers),
            f"📺 EPISODE {ep_num} — Detective: {det['name']} | Killer: {next(s['name'] for s in ep_data['suspects'] if s['is_killer'])}")
        row += 1
        
        # --- Model Roster Card ---
        style_subheader_row(ws, row, len(ep_headers), "🎬 INTRO ASSETS (Per-Episode)")
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"card_roster_ep{ep_num}", "Title Card",
            f"Model Roster — Episode {ep_num}",
            f"Detective: {det['name']}",
            "Auto-generated (PIL) — shows which model plays which character",
        ])
        row += 1
        
        # Roster details
        for sus in ep_data["suspects"]:
            role_note = " ⚔️ KILLER" if sus["is_killer"] else ""
            add_data_row(ws, row, [
                "", "", "",
                f"  {sus['character']}",
                f"Played by: {sus['name']}{role_note}",
                "",
            ])
            row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"card_specs_ep{ep_num}", "Title Card",
            f"Model Specs Card — Episode {ep_num}",
            "Exact model versions list",
            "Auto-generated (PIL)",
        ])
        row += 1
        
        # --- Scoring & Standings ---
        style_subheader_row(ws, row, len(ep_headers), "🏆 SCORING & STANDINGS (Per-Episode)")
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"card_scoring_ep{ep_num}", "Title Card",
            f"Scoring Breakdown — Episode {ep_num}",
            "Detective score, killer score, all innocent scores",
            "Auto-generated after episode transcript finalized",
        ])
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"card_standings_ep{ep_num}", "Title Card",
            f"Season Standings — Episode {ep_num}",
            "Running leaderboard after this episode",
            "Auto-generated — cumulative scores",
        ])
        row += 1
        
        # --- Episode Title ---
        add_data_row(ws, row, [
            "Not Started", f"card_title_ep{ep_num}", "Title Card",
            f"Episode Title Card — Episode {ep_num}",
            f"S1E{ep_num}: The Tomb of Amenhotep",
            "Auto-generated with episode subtitle if applicable",
        ])
        row += 1
        
        # --- TTS Audio ---
        style_subheader_row(ws, row, len(ep_headers), "🎙️ TTS AUDIO (Per-Episode)")
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"tts_narrator_ep{ep_num}", "TTS",
            f"Narrator Audio — Episode {ep_num}",
            "All narrator lines from transcript",
            "ElevenLabs — George voice",
        ])
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"tts_think_ep{ep_num}", "TTS",
            f"Model THINK Audio — Episode {ep_num}",
            "All model thinking lines — routed to model's permanent voice",
            "ElevenLabs — 8 model voices",
        ])
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"tts_speak_ep{ep_num}", "TTS",
            f"Character SPEAK Audio — Episode {ep_num}",
            "All character dialogue — routed to character's season voice",
            "ElevenLabs — 8 character voices",
        ])
        row += 1
        
        # --- Post-Production Overlays ---
        style_subheader_row(ws, row, len(ep_headers), "🎨 POST-PRODUCTION OVERLAYS (Per-Episode)")
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"overlay_lowerthirds_ep{ep_num}", "Overlay",
            f"Lower Thirds — Episode {ep_num}",
            "Colored bars: model icon + character name + model name per shot",
            f"Auto-generated based on Ep {ep_num} rotation",
        ])
        row += 1
        
        add_data_row(ws, row, [
            "Not Started", f"overlay_captions_ep{ep_num}", "Overlay",
            f"Dialogue Captions — Episode {ep_num}",
            "Colored text captions for all THINK and SPEAK sections",
            "Model brand colors for THINK, white/character for SPEAK",
        ])
        row += 1
        
        add_validation(ws, "A", 2, row)
        ws.freeze_panes = "A2"
    
    # =================================================================
    # Save
    # =================================================================
    wb.save(output_path)
    
    total_stills = sum(1 for r in ws1.iter_rows(min_row=2) if r[2].value and r[2].value != "")
    total_vids = sum(1 for r in ws2.iter_rows(min_row=2) if r[2].value and r[2].value != "")
    
    print(f"Built: {output_path}")
    print(f"Tabs: Reusable Stills | Reusable Videos | Episodes 1-8")
    print(f"Total reusable stills: {total_stills}")
    print(f"Total reusable videos: {total_vids}")
    print(f"Per-episode tabs: 8")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python build_tracker_xlsx.py <story_config.json> [output.xlsx]")
        sys.exit(1)
    
    with open(sys.argv[1]) as f:
        story = json.load(f)
    
    # Load rotation
    rotation_path = Path(sys.argv[1]).parent / "rotation.json"
    with open(rotation_path) as f:
        rotation = json.load(f)
    
    output = sys.argv[2] if len(sys.argv) > 2 else "s01_asset_tracker.xlsx"
    build_workbook(story, rotation, output)
